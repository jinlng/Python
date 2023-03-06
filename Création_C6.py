# update a version of prep_C6 specifically for création d'artère
# Import the openpyxl library for working with Excel files
import openpyxl
# Import the l93_to_wgs84 function from the coordinates_converter.py file
from coordinates_converter import l93_to_wgs84

# Define the path to the input Excel file
wbk_name = "C:/Users/liang.jingyi/Documents/CALCUL DE CHARGE/... ./Tableau CAP FT.xlsx"

# Define the commune name and INSEE code
commune = "St-André-de-Messei"
insee = '61362'
adresse = "Lieu dit Maudouet / l'Être au Moine"

# Modification of the types
mapping_dict = {
    'FR07': 'FL7',
    'FR 07': 'FL7',
    'FR 7': 'FL7',
    'FR08': 'FL8',
    'FR 08': 'FL8',
    'FR 8': 'FL8',
    'FR7_CREATION': 'FL7',
    'FR8_CREATION': 'FL8',
    'CREATION_FC7 MAX': 'FC7 MAX',
    'CREATION_FC7': 'FC7 MAX',
    'CREATION_FC8 MAX': 'FC8 MAX',
    'CREATION_FC8': 'FC8 MAX',
    'CREATION_FR7': 'FL7',
    'CREATION_FR8': 'FL8',
    'CREATION_FL7': 'FL7',
    'CREATION_FL8': 'FL8',
    'CREATION_MI8': 'M28',
    'CREATION_MI7': 'M27',
    'MI8': 'M28',
    'MI7': 'M27',
    'BH6': 'BH6 S30',
    'BH7': 'BH7 S30',
    'MH7': 'MH7 S30',
    'BH8': 'BH8 S30',
    'MH8': 'MH8 S30',
    'MR0': 'M20',
    'BETON': 'EDF'
}
# Define the name of the worksheet to work with
sheet_name = "Saisies terrain"

# Load the workbook and get the active worksheet
try:
    wb = openpyxl.load_workbook(wbk_name)
    ws = wb[sheet_name]
except Exception as e:
    print(f"Error loading workbook or worksheet: {e}")
    # You could add additional error handling here if needed, such as exiting the script

# Get the last row number in the worksheet
last_row = ws.max_row

# Convert Lambert 93 coordinates to WGS84 (latitude, longitude) and update the worksheet
for i in range(9, last_row + 1):
    value = ws.cell(row=i, column=19).value
    if value in ['L1092-11', 'L1092-13', 'L1092-14']:
        value += '-A'
    elif value == 'L1092-15':
        value = 'L1092-15-S'
    elif value == 'L1092-12-P':
        value = 'L1092-12-A'
    if value != ws.cell(row=i, column=19).value:
        ws.cell(row=i, column=19).value = value
        ws.cell(row=i, column=19).font = openpyxl.styles.Font(bold=True)

    if ws.cell(row=i, column=1).value and isinstance(ws.cell(row=i, column=4).value, float):
        # Get the Lambert 93 coordinates from the fourth and fifth columns of the current row
        x = ws.cell(row=i, column=4).value
        y = ws.cell(row=i, column=5).value
        # Convert the Lambert 93 coordinates to WGS84 (latitude, longitude) using the l93_to_wgs84 function
        lat, lon = l93_to_wgs84(x, y)
        # Update the fourth and fifth columns of the current row with the converted WGS84 coordinates
        ws.cell(row=i, column=4).value = lat
        ws.cell(row=i, column=5).value = lon

# Fill in specific values and update the worksheet
for i in range(9, last_row+1):
    if ws.cell(row=i, column=1).value is not None:
        # Fill in preset values in each line
        ws.cell(row=i, column=3).value = adresse
        for col, value in {
            15: 'TER',
            23: '0',
            24: '15',
            26: 'Oui',
            27: 'Non',
            35: 'Non',
        }.items():
            ws.cell(row=i, column=col).value = value
        # Replace type names
        cell_value = ws.cell(row=i, column=2).value
        if cell_value in mapping_dict:
            ws.cell(row=i, column=2).value = mapping_dict[cell_value]
        # if anything wrong, assign it as non-usable
        if ws.cell(row=i, column=6).value in ['FEN', 'EPA', 'CHO', 'AUT']:
            ws.cell(row=i, column=14).value = 'Non'
            for j in range(7, 14):
                ws.cell(row=i, column=j).value = 'Oui'
        if ws.cell(row=i, column=13).value == 'Oui' or not ws.cell(row=i, column=6).value:
            for j in range(6, 15):
                ws.cell(row=i, column=j).value = 'Oui'
        # check for EDF pillar and potelet
        if 'BT' in str(ws.cell(row=i, column=1).value):
            ws.cell(row=i, column=2).value = 'EDF'
        elif 'Ab' in str(ws.cell(row=i, column=1).value):
            ws.cell(row=i, column=2).value = 'POT MIN'
        else:
            continue


# Fill in specific cases
ws.cell(row=1, column=3).value = ''
ws.cell(row=1, column=6).value = ''
ws.cell(row=2, column=3).value = 'ORANGE'
ws.cell(row=2, column=6).value = 'SADE TELECOM'
ws.cell(row=4, column=4).value = 'Oui'
ws.cell(row=4, column=7).value = 'Oui'
ws.cell(row=6, column=4).value = 'A1'
ws.cell(row=6, column=5).value = 'B1'
ws.cell(row=3, column=3).value = commune
ws.cell(row=3, column=7).value = insee

wb.save(wbk_name)
wb.close