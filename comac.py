"""
Title: Export Comac Prep
Description:
    This code is designed to modify multiple files within a folder, specifically those named with the keyword 'ExportComac'.
    It updates specific columns in each file based on certain conditions, filling them with either 'D1' or 'D2' values.
    (Since we no longer calculer those in 'D3', but if we do in the future, this code is prepared for 'D3' as well.)
    However, there is a unknown bug in the code where each file needs to be individually saved before it can be accepted by apcom,
    and this issue has not been resolved yet. Please be aware of this limitation while using the code.
Author: Jingyi LIANG
Date: May 8, 2023
License: This code is the property of Jingyi LIANG. Unauthorized use or distribution is strictly prohibited.
"""

import os
import openpyxl

# Define the folder directory containing the files to modify
folder_path = "C:/Users/liang.jingyi/Documents/APCOM/Prep"

# Define the keyword used to filter the files to modify
keyword = "ExportComac"


def export_comac_prep(folder_path, keyword):
    # Recursively iterate through all directories and files in the folder
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for filename in filenames:
            # Check if the filename contains the keyword
            if keyword in filename:
                # Load the Excel workbook and active sheet
                wb = openpyxl.load_workbook(os.path.join(dirpath, filename))
                ws = wb.active

                # Get the last row in the sheet
                last_empty_row = ws.max_row + 1

                # Iterate through each row in the sheet, starting at row 4
                for i in range(4, last_empty_row):
                    if ws.cell(row=i, column=1).value is not None:
                        # If the value in column 47 is 0, set the value in column 46 to 'D3'
                        if ws.cell(row=i, column=47).value == 0:
                            ws.cell(row=i, column=46).value = 'D3'
                        # If the value in column 47 is greater than 0 and 'L1092-15' is in the value in column 41,
                        # set the value in column 46 to 'D1'
                        elif ws.cell(row=i, column=47).value > 0 and 'L1092-15' in ws.cell(row=i, column=41).value:
                            ws.cell(row=i, column=46).value = 'D1'
                        # If the value in column 46 is None, set it to 'D2'
                        elif ws.cell(row=i, column=46).value is None:
                            ws.cell(row=i, column=46).value = 'D2'
                    else:
                        pass

                # Save the changes to the workbook
                wb.save(os.path.join(dirpath, filename))

        for dirname in dirnames:
            # Recursively call the function on subdirectories
            export_comac_prep(os.path.join(dirpath, dirname), keyword)


# Call the function with the folder directory and keyword
export_comac_prep(folder_path, keyword)

