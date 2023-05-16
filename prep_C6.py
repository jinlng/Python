"""
Title: prep_C6 for calcul de charges CAP FT
Description:
    This code is designed to update Annex C6, initially with preset values.
    However, it is important to note that each "poteau" (supporting post) will need to be determined on a case-by-case basis afterwards.
    Additionally, the address details and path will need to be modified accordingly.
Author: Jingyi LIANG
Date: May 8, 2023
License: This code is the property of Jingyi LIANG. Unauthorized use or distribution is strictly prohibited.
"""

# Import the openpyxl library for working with Excel files
import openpyxl
# Import the l93_to_wgs84 function from the coordinates_converter.py file
from coordinates_converter import l93_to_wgs84

# Define the path to the input Excel file
wbk_name = "C:/Users/liang.jingyi/Documents/CALCUL DE CHARGE/LOA/29102/Tableau CAP FT.xlsx"

# Define the commune name and INSEE code
commune = "Lonlay-l'Abbaye"
insee = '61232'
adresse = "Lieu dit la Douardière"

mapping_dict = {
    'Bois': 'BS7',
    'FR07': 'FL7',
    'FR 07': 'FL7',
    'FR 7': 'FL7',
    'FR08': 'FL8',
    'FR 08': 'FL8',
    'FR 8': 'FL8',
    'FR7_CREATION': 'FL7',
    'FR8_CREATION': 'FL8',
    'CREATION_FC7 MAX': 'FC7 MAX',
    'CREATION_FC7': 'FC7 MAX',
    'CREATION_FC8 MAX': 'FC8 MAX',
    'CREATION_FC8': 'FC8 MAX',
    'CREATION_FR7': 'FL7',
    'CREATION_FR8': 'FL8',
    'CREATION_FL7': 'FL7',
    'CREATION_FL8': 'FL8',
    'CREATION_MI8': 'M28',
    'CREATION_MI7': 'M27',
    'MI8': 'M28',
    'MI7': 'M27',
    'BH6': 'BH6 S30',
    'BH7': 'BH7 S30',
    'MH7': 'MH7 S30',
    'BH8': 'BH8 S30',
    'MH8': 'MH8 S30',
    'MR0': 'M20',
    'BETON': 'EDF',
    'ERDF': 'EDF',
    'BT': 'EDF',
    'Ab': 'POT MIN'
}
# Define the name of the worksheet to work with
sheet_name = "Saisies terrain"

# Load the workbook and get the active worksheet
try:
    wb = openpyxl.load_workbook(wbk_name)
    ws = wb[sheet_name]
except Exception as e:
    print(f"Error loading workbook or worksheet: {e}")
    # You could add additional error handling here if needed, such as exiting the script

# Get the last row number in the worksheet
last_row = ws.max_row

# Convert Lambert 93 coordinates to WGS84 (latitude, longitude) and update the worksheet
for i in range(9, last_row+1):
    value = ws.cell(row=i, column=19).value
    if value in ['L1092-11', 'L1092-13', 'L1092-14']:
        value += '-A'
    elif value == 'L1092-15':
        value = 'L1092-15-S'
    elif value == 'L1092-12-P':
        value = 'L1092-12-A'
    if value != ws.cell(row=i, column=19).value:
        ws.cell(row=i, column=19).value = value
        ws.cell(row=i, column=19).font = openpyxl.styles.Font(bold=True)
    if ws.cell(row=i, column=1).value is not None and isinstance(ws.cell(row=i, column=4).value, float):
        # Convert the Lambert 93 coordinates to WGS84 (latitude, longitude) using the l93_to_wgs84 function
        lat, lon = l93_to_wgs84(ws.cell(row=i, column=4).value, ws.cell(row=i, column=5).value)
        # Update the fourth and fifth columns of the current row with the converted WGS84 coordinates
        ws.cell(row=i, column=4).value = lat
        ws.cell(row=i, column=5).value = lon

# Fill in specific values and update the worksheet
for i in range(9, last_row+1):
    # for every poteau line
    if ws.cell(row=i, column=1).value is not None:
        # Fill in preset values in each line
        ws.cell(row=i, column=3).value = adresse
        for col, value in {
            15: 'TER',
            23: '0',
            24: '15',
            26: 'Oui',
            27: 'Non',
            35: 'Oui',
            36: 'Non',
            37: '0',
            38: 'Non',
            39: '0',
            40: 'Non',
        }.items():
            ws.cell(row=i, column=col).value = value
        # Replace poteau type names
        cell_value = ws.cell(row=i, column=2).value
        if cell_value in mapping_dict:
            ws.cell(row=i, column=2).value = mapping_dict[cell_value]
        # if anything wrong, assign it as non-usable
        if ws.cell(row=i, column=6).value in ['FEN', 'EPA', 'CHO', 'AUT']:
            ws.cell(row=i, column=14).value = 'Non'
            for j in range(7, 14):
                ws.cell(row=i, column=j).value = 'Oui'
        if ws.cell(row=i, column=13).value == 'Oui' or not ws.cell(row=i, column=6).value:
            for j in range(6, 15):
                ws.cell(row=i, column=j).value = 'Oui'

    # fill temperature and flèche cases for every line
    if ws.cell(row=i, column=19).value is not None:
        for col, value in {
            23: '0',
            24: '15',
        }.items():
            ws.cell(row=i, column=col).value = value



# Fill in specific cases
ws.cell(row=1, column=3).value = ''
ws.cell(row=1, column=6).value = ''
ws.cell(row=2, column=3).value = 'ORANGE'
ws.cell(row=2, column=6).value = 'SADE TELECOM'
ws.cell(row=4, column=4).value = 'Oui'
ws.cell(row=4, column=7).value = 'Oui'
ws.cell(row=6, column=4).value = 'A1'
ws.cell(row=6, column=5).value = 'B1'
ws.cell(row=3, column=3).value = commune
ws.cell(row=3, column=7).value = insee

wb.save(wbk_name)
wb.close